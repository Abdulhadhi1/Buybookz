import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import os
import re

def parse_item(item):
    # Book Title
    title_tag = item.find('h4')
    title = title_tag.text.strip() if title_tag else ''
    
    # Image
    img_div = item.find('div', class_='product_list_img')
    img_url = img_div.get('data-src', '') if img_div else ''
    if not img_url and img_div:
        style = img_div.get('style', '')
        if 'url' in style:
            match = re.search(r'url\([\'"]?(.*?)[\'"]?\)', style)
            if match:
                img_url = match.group(1)
                
    # Price
    price_tag = item.find('font', class_='price')
    price = price_tag.text.strip() if price_tag else ''
    
    # MRP
    mrp_tag = item.find('font', class_='mrp')
    mrp = mrp_tag.text.strip() if mrp_tag else ''
    
    # Offer
    offer_tag = item.find('font', class_='offer')
    offer = offer_tag.text.strip() if offer_tag else ''
    
    # Author
    author_tag = item.find('h6', class_='author')
    author = ''
    if author_tag:
        author_link = author_tag.find('a')
        if author_link:
            author = author_link.text.strip()
        else:
            text = author_tag.text
            text = text.replace('Author', '').replace(':', '').strip()
            author = text
            
    # Publisher
    publisher_tag = item.find('h6', class_='publisher')
    publisher = ''
    if publisher_tag:
        publisher_link = publisher_tag.find('a')
        if publisher_link:
            publisher = publisher_link.text.strip()
        else:
            text = publisher_tag.text
            text = text.replace('Publisher', '').replace(':', '').strip()
            publisher = text
            
    # Pages
    pages_tag = item.find('h6', class_='pages')
    pages = ''
    if pages_tag:
        text = pages_tag.text
        text = text.replace('No. of pages', '').replace(':', '').strip()
        pages = text

    return {
        'Book Name': title,
        'Author': author,
        'Price': price,
        'MRP': mrp,
        'Discount': offer,
        'Publisher': publisher,
        'Pages': pages,
        'Image Link': img_url
    }

def main():
    books = []
    page = 1
    target_count = 500
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36'
    }
    
    print("==================================================")
    print("      COMMONFOLKS BOOK DETAILS SCRAPER            ")
    print("==================================================")
    print(f"Goal: Scrape at least {target_count} books details.")
    
    while len(books) < target_count:
        url = f'https://www.commonfolks.in/books?f[page]={page}&f[sort]=default&f[view]=list'
        print(f"\nScraping Page {page}: {url}")
        
        try:
            response = requests.get(url, headers=headers, timeout=15)
            if response.status_code != 200:
                print(f"Error: Status code {response.status_code}. Stopping.")
                break
                
            soup = BeautifulSoup(response.text, 'html.parser')
            items = soup.find_all('div', class_='item')
            
            if not items:
                print("No more books found or reached the end. Stopping.")
                break
                
            print(f"Found {len(items)} books on Page {page}.")
            
            for item in items:
                book_data = parse_item(item)
                # Only add if we have a title (valid book)
                if book_data['Book Name']:
                    books.append(book_data)
                    
            print(f"Total books gathered so far: {len(books)} / {target_count}")
            
            # Rate limiting delay
            time.sleep(0.5)
            page += 1
            
        except Exception as e:
            print(f"Exception occurred: {e}. Retrying after 5 seconds...")
            time.sleep(5)
            
    # Slice to exactly target count if we have more
    final_books = books[:target_count]
    print("\n==================================================")
    print(f"Scraping complete! Gathered {len(books)} books total.")
    print(f"Slicing to exactly {len(final_books)} books for output.")
    print("==================================================")
    
    df = pd.DataFrame(final_books)
    
    # Save CSV
    csv_file = 'CommonFolks_500_Books.csv'
    df.to_csv(csv_file, index=False, encoding='utf-8-sig')
    print(f"Saved CSV file: {csv_file}")
    
    # Save formatted Excel XLSX
    xlsx_file = 'CommonFolks_500_Books.xlsx'
    
    # Write to Excel with custom styling
    with pd.ExcelWriter(xlsx_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Books', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['Books']
        
        # Styles
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Navy Blue
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )
        
        # Apply header styling
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            
        # Apply data styling
        for row_num in range(2, len(final_books) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = thin_border
                
                # Check column and apply specific alignments
                col_name = df.columns[col_num - 1]
                if col_name in ['Price', 'MRP', 'Discount', 'Pages']:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_name == 'Image Link':
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    # Give it a subtle hyperlinked styling if it looks like a URL
                    if cell.value:
                        cell.font = Font(name="Calibri", size=10, color="0000FF", underline="single")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Auto-adjust column widths
        for col in worksheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            # Add padding
            worksheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)
            
        # Set Row Heights
        worksheet.row_dimensions[1].height = 26  # Header row
        for r in range(2, len(final_books) + 2):
            worksheet.row_dimensions[r].height = 20
            
    print(f"Saved styled Excel file: {xlsx_file}")
    print("==================================================")
    print("Done!")

if __name__ == '__main__':
    main()
