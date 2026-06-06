import requests
from bs4 import BeautifulSoup
import pandas as pd
import concurrent.futures
import time
import os
import re
from threading import Lock

# Locks for thread-safe operations
print_lock = Lock()
books_lock = Lock()

books = []
pages_scraped = 0
total_pages = 3015

def parse_item(item):
    # Book Title
    title_tag = item.find('h4')
    title = title_tag.text.strip() if title_tag else ''
    
    # Image
    img_div = item.find('div', class_='product_list_img')
    img_url = img_div.get('data-src', '') if img_div else ''
    if not img_url and img_div:
        style = img_div.get('style', '')
        if 'url' in style:
            match = re.search(r'url\([\'"]?(.*?)[\'"]?\)', style)
            if match:
                img_url = match.group(1)
                
    # Price
    price_tag = item.find('font', class_='price')
    price = price_tag.text.strip() if price_tag else ''
    
    # MRP
    mrp_tag = item.find('font', class_='mrp')
    mrp = mrp_tag.text.strip() if mrp_tag else ''
    
    # Offer
    offer_tag = item.find('font', class_='offer')
    offer = offer_tag.text.strip() if offer_tag else ''
    
    # Author
    author_tag = item.find('h6', class_='author')
    author = ''
    if author_tag:
        author_link = author_tag.find('a')
        if author_link:
            author = author_link.text.strip()
        else:
            text = author_tag.text
            text = text.replace('Author', '').replace(':', '').strip()
            author = text
            
    # Publisher
    publisher_tag = item.find('h6', class_='publisher')
    publisher = ''
    if publisher_tag:
        publisher_link = publisher_tag.find('a')
        if publisher_link:
            publisher = publisher_link.text.strip()
        else:
            text = publisher_tag.text
            text = text.replace('Publisher', '').replace(':', '').strip()
            publisher = text
            
    return {
        'Book Name': title,
        'Author': author,
        'Price': price,
        'MRP': mrp,
        'Discount': offer,
        'Publisher': publisher,
        'Image Link': img_url
    }

def scrape_page(page_num):
    global pages_scraped
    url = f'https://www.commonfolks.in/books?f[page]={page_num}&f[sort]=default&f[view]=list'
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36'
    }
    
    max_retries = 3
    for attempt in range(max_retries):
        try:
            response = requests.get(url, headers=headers, timeout=15)
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, 'html.parser')
                items = soup.find_all('div', class_='item')
                
                page_books = []
                for item in items:
                    book_data = parse_item(item)
                    if book_data['Book Name']:
                        page_books.append(book_data)
                
                with books_lock:
                    books.extend(page_books)
                    pages_scraped += 1
                    current_count = len(books)
                    current_pages = pages_scraped
                
                if current_pages % 50 == 0 or current_pages == total_pages:
                    with print_lock:
                        print(f"Progress: Scraped {current_pages}/{total_pages} pages ({round(current_pages/total_pages*100, 1)}%) | Gathered {current_count} books total.")
                
                return
            elif response.status_code == 429:
                with print_lock:
                    print(f"Rate limited on Page {page_num}. Sleeping for 5s... (Attempt {attempt+1}/{max_retries})")
                time.sleep(5)
            else:
                with print_lock:
                    print(f"Warning: Page {page_num} returned status code {response.status_code} (Attempt {attempt+1}/{max_retries})")
                time.sleep(1)
        except Exception as e:
            if attempt == max_retries - 1:
                with print_lock:
                    print(f"Error scraping Page {page_num} on final attempt: {e}")
            time.sleep(2)

def main():
    start_time = time.time()
    
    print("==================================================")
    print("      COMMONFOLKS FULL CATALOG BOOK SCRAPER       ")
    print("==================================================")
    print(f"Goal: Scrape ALL books from all {total_pages} pages.")
    print("Configured: 15 Concurrent Worker Threads.")
    print("Pages Column: Omitted as requested.")
    print("==================================================")
    
    pages = list(range(1, total_pages + 1))
    
    # ThreadPool execution
    with concurrent.futures.ThreadPoolExecutor(max_workers=15) as executor:
        executor.map(scrape_page, pages)
        
    duration = time.time() - start_time
    total_gathered = len(books)
    print("\n==================================================")
    print(f"Scraping complete in {round(duration, 2)} seconds (~{round(duration/60, 2)} minutes)!")
    print(f"Total books gathered: {total_gathered}")
    print("==================================================")
    
    # Structure data
    print("Structuring and sorting data by Book Name...")
    df = pd.DataFrame(books)
    # Sort books by title alphabetically
    df = df.sort_values(by='Book Name').reset_index(drop=True)
    
    # Save CSV
    csv_file = 'CommonFolks_All_Books.csv'
    df.to_csv(csv_file, index=False, encoding='utf-8-sig')
    print(f"Saved CSV file: {csv_file}")
    
    # Save styled Excel XLSX
    xlsx_file = 'CommonFolks_All_Books.xlsx'
    print("Saving and styling Excel workbook...")
    
    with pd.ExcelWriter(xlsx_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='All Books', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['All Books']
        
        # Styles
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Navy Blue
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )
        
        # Apply header styling
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            
        # Apply data styling
        print("Applying cell alignments and borders...")
        for row_num in range(2, total_gathered + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = thin_border
                
                # Check column and apply specific alignments
                col_name = df.columns[col_num - 1]
                if col_name in ['Price', 'MRP', 'Discount']:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_name == 'Image Link':
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    if cell.value:
                        cell.font = Font(name="Calibri", size=10, color="0000FF", underline="single")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Auto-adjust column widths
        print("Calculating and adjusting column widths...")
        for col in worksheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            # Sample max length from headers + data (we sample every 5th row to speed up column width calculation for 80k+ rows)
            sampled_cells = [col[0]] + list(col[1::5])
            for cell in sampled_cells:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            worksheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)
            
        # Set Row Heights
        worksheet.row_dimensions[1].height = 26  # Header row
        # Apply standard height to data rows
        print("Finalizing sheet dimensions...")
        
    print(f"Saved styled Excel file: {xlsx_file}")
    print("==================================================")
    print("All tasks completed successfully!")

if __name__ == '__main__':
    main()
